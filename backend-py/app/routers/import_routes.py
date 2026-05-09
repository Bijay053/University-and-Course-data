"""/api/import endpoints — Excel (XLSX) bulk import for staged courses."""
from __future__ import annotations

import io
import re
import uuid
import zipfile
from datetime import datetime, timezone
from typing import Annotated, Any

from fastapi import APIRouter, Depends, Form, HTTPException, UploadFile, File, status
from openpyxl import load_workbook
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_current_user, get_db
from app.models import University
from app.models.import_job import ImportJob
from app.models.scraped_course import ScrapedCourse

router = APIRouter(prefix="/import", tags=["import"])

MAX_BYTES = 10 * 1024 * 1024            # 10 MB compressed upload cap
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024   # 100 MB total inflated cap
MAX_ENTRIES = 200                       # XLSX usually has < 50 parts
MAX_COMPRESSION_RATIO = 200             # any single entry inflating > 200x is suspicious


def _validate_xlsx_archive(raw: bytes) -> None:
    """Reject zip-bombs before handing the file to openpyxl.

    XLSX is a ZIP container. Without these checks an attacker could upload a
    small file whose internal XML inflates to gigabytes during parsing,
    exhausting CPU/RAM. We inspect the central directory only — no entry data
    is decompressed here.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            infos = zf.infolist()
            if len(infos) > MAX_ENTRIES:
                raise HTTPException(
                    status_code=400,
                    detail={"error": f"Workbook has too many internal parts ({len(infos)})."},
                )
            total_uncompressed = 0
            for info in infos:
                total_uncompressed += info.file_size
                if info.compress_size > 0 and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO:
                    raise HTTPException(
                        status_code=400,
                        detail={"error": "Workbook entry has suspicious compression ratio (possible zip bomb)."},
                    )
            if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise HTTPException(
                    status_code=400,
                    detail={"error": f"Workbook would inflate above {MAX_UNCOMPRESSED_BYTES // (1024 * 1024)} MB."},
                )
    except zipfile.BadZipFile as exc:
        raise HTTPException(
            status_code=400,
            detail={"error": "File is not a valid .xlsx workbook."},
        ) from exc


# ─── /api/import/history ────────────────────────────────────────────────────
@router.get("/history")
async def import_history(db: Annotated[AsyncSession, Depends(get_db)]) -> list:
    """Return list of past import jobs. Empty list if table doesn't exist."""
    try:
        rows = await db.execute(text(
            "SELECT id, file_name, status, total_rows, imported_rows, "
            "skipped_rows, NULL::int AS error_rows, created_at, completed_at "
            "FROM import_jobs ORDER BY created_at DESC LIMIT 50"
        ))
        return [
            {
                "id": r[0],
                "fileName": r[1],
                "status": r[2],
                "totalRows": r[3],
                "importedRows": r[4],
                "skippedRows": r[5],
                "errorRows": r[6],
                "createdAt": r[7].isoformat() if r[7] else None,
                "completedAt": r[8].isoformat() if r[8] else None,
            }
            for r in rows.fetchall()
        ]
    except Exception:
        return []


# ─── Column mapping ─────────────────────────────────────────────────────────
# Maps a normalised header (lowercase, alphanumeric only) → ScrapedCourse attr.
_COLUMN_MAP: dict[str, str] = {
    "coursename": "course_name",
    "name": "course_name",
    "category": "category",
    "subcategory": "sub_category",
    "degreelevel": "degree_level",
    "level": "degree_level",
    "duration": "duration",
    "durationterm": "duration_term",
    "studymode": "study_mode",
    "mode": "study_mode",
    "studyload": "study_load",
    "intakemonth": "intake_months",
    "intakemonths": "intake_months",
    "intake": "intake_months",
    "internationalfee": "international_fee",
    "fee": "international_fee",
    "feeterm": "fee_term",
    "feeyear": "fee_year",
    "currency": "currency",
    "ieltsoverall": "ielts_overall",
    "ieltslistening": "ielts_listening",
    "ieltsspeaking": "ielts_speaking",
    "ieltswriting": "ielts_writing",
    "ieltsreading": "ielts_reading",
    "pteoverall": "pte_overall",
    "ptelistening": "pte_listening",
    "ptespeaking": "pte_speaking",
    "ptewriting": "pte_writing",
    "ptereading": "pte_reading",
    "toefloverall": "toefl_overall",
    "academiclevel": "academic_level",
    "academicscore": "academic_score",
    "academiccountry": "academic_country",
    "scholarship": "scholarship",
    "coursewebsite": "course_website",
    "website": "course_website",
    "url": "course_website",
    "courseurl": "course_website",
    "courselocation": "course_location",
    "location": "course_location",
    "campus": "course_location",
    "description": "description",
    "otherrequirement": "other_requirement",
    "language": "language",
    "cricoscode": "cricos_code",
    "cricos": "cricos_code",
}

_FLOAT_FIELDS = {
    "duration", "international_fee", "academic_score",
    "ielts_overall", "ielts_listening", "ielts_speaking", "ielts_writing", "ielts_reading",
    "pte_overall", "pte_listening", "pte_speaking", "pte_writing", "pte_reading",
    "toefl_overall",
}
_INT_FIELDS = {"fee_year"}
_LIST_FIELDS = {"intake_months"}

_MONTH_NAMES = {
    "jan": "January", "january": "January",
    "feb": "February", "february": "February",
    "mar": "March", "march": "March",
    "apr": "April", "april": "April",
    "may": "May",
    "jun": "June", "june": "June",
    "jul": "July", "july": "July",
    "aug": "August", "august": "August",
    "sep": "September", "sept": "September", "september": "September",
    "oct": "October", "october": "October",
    "nov": "November", "november": "November",
    "dec": "December", "december": "December",
}


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(h).strip().lower())


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Strip currency / commas / "AUD" etc.
    cleaned = re.sub(r"[^\d.\-]", "", s)
    if not cleaned or cleaned in {"-", ".", "-."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_int(v: Any) -> int | None:
    f = _to_float(v)
    return int(f) if f is not None else None


def _to_intake_months(v: Any) -> list[str] | None:
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    parts = re.split(r"[,;/|]+", s)
    out: list[str] = []
    for p in parts:
        key = p.strip().lower()
        key = re.sub(r"[^a-z]", "", key)
        if not key:
            continue
        if key in _MONTH_NAMES:
            month = _MONTH_NAMES[key]
            if month not in out:
                out.append(month)
        else:
            # Unknown token — keep raw (truncated) so reviewer can see it.
            kept = p.strip()[:32]
            if kept and kept not in out:
                out.append(kept)
    return out or None


def _coerce(field: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    if field in _FLOAT_FIELDS:
        return _to_float(value)
    if field in _INT_FIELDS:
        return _to_int(value)
    if field in _LIST_FIELDS:
        return _to_intake_months(value)
    if isinstance(value, str):
        return value.strip() or None
    return value


# ─── /api/import/excel ──────────────────────────────────────────────────────
@router.post("/excel")
async def import_excel(
    db: Annotated[AsyncSession, Depends(get_db)],
    _user: Annotated[dict, Depends(get_current_user)],
    file: UploadFile = File(...),
    universityId: str | None = Form(None),
    universityName: str | None = Form(None),
    universityCountry: str | None = Form(None),
    universityCity: str | None = Form(None),
) -> dict[str, Any]:
    """Bulk-import course rows from an XLSX file into ``scraped_courses``.

    The frontend sends either ``universityId`` (existing uni) or
    ``universityName`` + optional country/city (create-on-the-fly).
    Each row becomes a ``scraped_courses`` row with ``status='pending'`` and
    ``auto_publish_status='pending_review'`` so it surfaces in the normal
    review queue. Duplicates within the same university (case-insensitive
    course name) are skipped, not raised.
    """
    fname = file.filename or "upload.xlsx"
    if not fname.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail={"error": "File must be an .xlsx Excel workbook."},
        )
    raw = await file.read(MAX_BYTES + 1)
    if not raw:
        raise HTTPException(status_code=400, detail={"error": "Empty file."})
    if len(raw) > MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail={"error": f"File exceeds {MAX_BYTES // (1024 * 1024)} MB limit."},
        )
    _validate_xlsx_archive(raw)

    # ── Resolve target university ──────────────────────────────────────────
    uni: University | None = None
    if universityId:
        try:
            uni_id_int = int(universityId)
        except ValueError:
            raise HTTPException(
                status_code=400, detail={"error": "universityId must be an integer."}
            ) from None
        uni = await db.get(University, uni_id_int)
        if uni is None:
            raise HTTPException(
                status_code=404,
                detail={"error": f"University id={uni_id_int} not found."},
            )
    elif universityName and universityName.strip():
        name_clean = universityName.strip()
        existing = (
            await db.execute(
                select(University).where(func.lower(University.name) == name_clean.lower())
            )
        ).scalar_one_or_none()
        if existing is not None:
            uni = existing
        else:
            uni = University(
                name=name_clean,
                country=(universityCountry or "").strip() or "Unknown",
                city=(universityCity or "").strip() or "Unknown",
            )
            db.add(uni)
            await db.flush()  # populate uni.id without committing yet
    else:
        raise HTTPException(
            status_code=400,
            detail={"error": "Provide either universityId or universityName."},
        )

    # ── Parse workbook ─────────────────────────────────────────────────────
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # InvalidFileException, zipfile.BadZipFile, …
        raise HTTPException(
            status_code=400,
            detail={"error": f"Could not read Excel file: {exc.__class__.__name__}"},
        ) from exc

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=400, detail={"error": "Workbook has no active sheet."}
        )

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(
            status_code=400, detail={"error": "Sheet is empty."}
        ) from None

    # Build header → field index map. Unknown columns are silently ignored.
    col_to_field: dict[int, str] = {}
    for idx, raw_header in enumerate(header_row):
        norm = _norm_header(raw_header)
        if norm in _COLUMN_MAP:
            col_to_field[idx] = _COLUMN_MAP[norm]
    if "course_name" not in col_to_field.values():
        raise HTTPException(
            status_code=400,
            detail={"error": "Sheet must include a 'Course Name' column."},
        )

    job_id = f"excel-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"

    # Load existing names once for dedupe (case-insensitive).
    existing_names = {
        n.lower()
        for (n,) in (
            await db.execute(
                select(ScrapedCourse.course_name).where(
                    ScrapedCourse.university_id == uni.id
                )
            )
        ).all()
        if n
    }

    total_rows = 0
    imported = 0
    skipped = 0
    errors: list[str] = []

    for line_no, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or cell == "" for cell in row):
            continue
        total_rows += 1

        payload: dict[str, Any] = {}
        for idx, field in col_to_field.items():
            if idx >= len(row):
                continue
            payload[field] = _coerce(field, row[idx])

        course_name = payload.get("course_name")
        if not course_name or not str(course_name).strip():
            errors.append(f"Row {line_no}: missing course name — skipped.")
            skipped += 1
            continue
        course_name = str(course_name).strip()
        payload["course_name"] = course_name

        if course_name.lower() in existing_names:
            skipped += 1
            continue
        existing_names.add(course_name.lower())

        try:
            db.add(ScrapedCourse(
                scrape_job_id=job_id,
                university_id=uni.id,
                status="pending",
                auto_publish_status="pending_review",
                eligibility_status="unknown",
                **payload,
            ))
            imported += 1
        except Exception as exc:  # type-coercion / column mismatch
            errors.append(f"Row {line_no}: {exc.__class__.__name__}: {exc}")
            skipped += 1

    # Record the import job row for /api/import/history.
    db.add(ImportJob(
        university_id=uni.id,
        university_name=uni.name,
        file_name=fname,
        status="completed" if not errors else "completed_with_errors",
        total_rows=total_rows,
        imported_rows=imported,
        skipped_rows=skipped,
        error_message=("; ".join(errors[:5]) if errors else None),
        completed_at=datetime.now(timezone.utc),
    ))

    try:
        await db.commit()
    except Exception as exc:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail={"error": f"Database commit failed: {exc.__class__.__name__}"},
        ) from exc

    return {
        "universityName": uni.name,
        "totalRows": total_rows,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
    }
